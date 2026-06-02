import base64
import io
from copy import deepcopy
from typing import BinaryIO

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils.exceptions import InvalidFileException

from src.exceptions import InvalidFileExtension
from src.logger import logger
from src.schemas import CellData, ImageData, TemplatePreviewResponse, TemplateSheet


class TemplateManager:
    """Gerencia o template Excel: abertura, clonagem de abas e salvamento.

    Carrega o workbook e extrai imagens no construtor — não requer
    chamada explícita a ``load()``.
    """

    def __init__(self, file: BinaryIO):
        self.file = file
        self.wb, self.ws_template, self.extracted_images = self._load(file)

    @property
    def worksheets(self):
        """Abas do workbook carregado."""
        return self.wb.worksheets

    def _load(self, file: BinaryIO) -> tuple:
        """Abre o workbook, extrai imagens e retorna (wb, ws_template, images)."""
        try:
            wb = load_workbook(file)
            ws_template = wb.active
            images: list[dict] = []
            ws_images: list = getattr(ws_template, "_images", [])
            for img in ws_images:
                result = _read_image_data(img)
                if result:
                    img_bytes, anchor = result
                    images.append({"bytes": img_bytes, "anchor": anchor})
                else:
                    logger.warning("Nao foi possivel ler uma imagem do template")
            setattr(ws_template, "_images", [])  # noqa: B010
            return wb, ws_template, images
        except InvalidFileException as e:
            logger.exception("Erro ao abrir template Excel")
            raise InvalidFileExtension(
                f"O arquivo de template nao e um Excel (.xlsx) valido: {e}"
            ) from e

    def clone_worksheet(self, title: str):
        new_ws = self.wb.copy_worksheet(self.ws_template)
        new_ws.title = title
        for img_info in self.extracted_images:
            try:
                new_img = OpenpyxlImage(io.BytesIO(img_info["bytes"]))
                # deepcopy do anchor é frágil — depende de implementação
                # interna do openpyxl, mas é necessário para clonar a
                # posição exata da imagem no novo worksheet.
                new_img.anchor = deepcopy(img_info["anchor"])
                new_ws.add_image(new_img)
            except (OSError, ValueError, AttributeError) as e:
                logger.warning("Nao foi possivel injetar imagem na aba %s: %s", title, e)
        return new_ws

    def save_to_stream(self, stream: BinaryIO):
        if self.ws_template:
            self.wb.remove(self.ws_template)
        self.wb.save(stream)
        logger.info("Relatorio salvo no stream")


def _read_image_data(img) -> tuple | None:
    """Extrai bytes e anchor de uma imagem do openpyxl.

    Retorna ``(bytes, anchor)`` ou ``None`` se a leitura falhar.
    """
    try:
        return img._data(), img.anchor
    except (OSError, ValueError, AttributeError):
        return None


_BORDER_STYLE_MAP = {
    "thin": "1px solid",
    "hair": "1px solid",
    "medium": "2px solid",
    "thick": "3px solid",
    "double": "3px double",
    "dashed": "1px dashed",
    "dotted": "1px dotted",
    "mediumDashed": "2px dashed",
    "dashDot": "1px dashed",
    "mediumDashDot": "2px dashed",
    "dashDotDot": "1px dotted",
    "mediumDashDotDot": "2px dotted",
    "slantDashDot": "1px dashed",
}


def _cell_style(cell) -> dict[str, str]:
    style: dict[str, str] = {}

    try:
        fill = cell.fill
        if fill and fill.patternType and fill.fgColor:
            rgb = fill.fgColor.rgb
            if rgb and rgb not in ("00000000", "0"):
                style["background-color"] = f"#{rgb[-6:]}"
    except (AttributeError, TypeError, ValueError):
        pass

    try:
        font = cell.font
        if font:
            if font.bold:
                style["font-weight"] = "bold"
            if font.italic:
                style["font-style"] = "italic"
            if font.underline and font.underline != "none":
                style["text-decoration"] = "underline"
            if font.size:
                style["font-size"] = f"{font.size:g}pt"
            if font.name:
                style["font-family"] = font.name
            if font.color and font.color.rgb:
                rgb = font.color.rgb
                if rgb and rgb not in ("00000000", "0"):
                    style["color"] = f"#{rgb[-6:]}"
    except (AttributeError, TypeError, ValueError):
        pass

    try:
        align = cell.alignment
        if align:
            if align.horizontal:
                style["text-align"] = align.horizontal
            if align.vertical:
                style["vertical-align"] = align.vertical
            if align.wrapText:
                style["white-space"] = "pre-wrap"
    except (AttributeError, TypeError, ValueError):
        pass

    try:
        border = cell.border
        if border:
            for side_name, css_prop in [
                ("left", "border-left"),
                ("right", "border-right"),
                ("top", "border-top"),
                ("bottom", "border-bottom"),
            ]:
                side = getattr(border, side_name, None)
                if side and side.style:
                    base = _BORDER_STYLE_MAP.get(side.style)
                    if base:
                        color = "#000000"
                        try:
                            if side.color and side.color.rgb:
                                rgb = side.color.rgb
                                if rgb and rgb not in ("00000000", "0"):
                                    color = f"#{rgb[-6:]}"
                        except (AttributeError, TypeError, ValueError):
                            pass
                        style[css_prop] = f"{base} {color}"
    except Exception:
        pass

    return style


def preview_template(file_bytes: bytes, filename: str) -> TemplatePreviewResponse:
    if not filename.lower().endswith(".xlsx"):
        raise InvalidFileExtension("Template deve ser .xlsx")

    buffer = io.BytesIO(file_bytes)
    tm = TemplateManager(buffer)

    all_sheets = []
    for ws in tm.worksheets:
        cells = []
        for row in ws.iter_rows():
            for cell in row:
                font_info = None
                if cell.font:
                    font_info = {"bold": cell.font.bold, "size": cell.font.size}
                cells.append(
                    CellData(
                        coord=cell.coordinate or "",
                        row=cell.row or 0,
                        col=cell.column or 0,
                        value=str(cell.value) if cell.value is not None else None,
                        font=font_info,
                        style=_cell_style(cell),
                    )
                )

        merged = []
        for r in ws.merged_cells.ranges:
            merged.append({
                "min_col": r.min_col,
                "max_col": r.max_col,
                "min_row": r.min_row,
                "max_row": r.max_row,
            })

        col_widths = {}
        for col_letter, dim in ws.column_dimensions.items():
            if dim.width:
                col_widths[col_letter] = dim.width

        row_heights = {}
        for row_num, dim in ws.row_dimensions.items():
            if dim.height:
                row_heights[row_num] = dim.height

        images = []
        for img in getattr(ws, "_images", []):
            result = _read_image_data(img)
            if result is None:
                logger.warning("Failed to extract image")
                continue
            img_bytes, _anchor = result
            b64 = base64.b64encode(img_bytes).decode()
            position = {}
            if hasattr(img.anchor, "_from"):
                position["col"] = img.anchor._from.col
                position["row"] = img.anchor._from.row
                position["colOff"] = img.anchor._from.colOff
                position["rowOff"] = img.anchor._from.rowOff
            images.append(ImageData(b64=f"data:image/png;base64,{b64}", position=position))

        all_sheets.append(TemplateSheet(
            name=ws.title,
            cells=cells,
            images=images,
            merged=merged,
            col_widths=col_widths or None,
            row_heights=row_heights or None,
        ))

    logger.info("Preview template: %d sheets from %s", len(all_sheets), filename)
    return TemplatePreviewResponse(sheets=all_sheets)
