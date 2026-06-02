import io

import pandas as pd
from openpyxl import Workbook

from src.excel import preview_source, preview_template


def _create_source_xlsx() -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        pd.DataFrame({"Data": ["2026-01-01"], "Atividade": ["Escavacao"]}).to_excel(
            writer, sheet_name="Obra", index=False
        )
    buffer.seek(0)
    return buffer.read()


def _create_source_xlsx_with_acronyms() -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "Data": ["2026-01-01", "2026-01-02"],
                "Serviço": ["Pintura LED externa", "Troca de PCD e BHLS"],
            }
        ).to_excel(writer, sheet_name="Obra", index=False)
    buffer.seek(0)
    return buffer.read()


def _create_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "RDO Header"
    ws["B3"] = "Servicos:"
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def test_preview_source():
    data = _create_source_xlsx()
    result = preview_source(data, "medicao.xlsx")
    assert len(result.sheets) == 1
    assert result.sheets[0].name == "Obra"
    assert "Atividade" in result.sheets[0].columns


def test_preview_source_suggests_acronyms():
    data = _create_source_xlsx_with_acronyms()
    result = preview_source(data, "medicao.xlsx")
    assert "LED" in result.suggestedAcronyms
    assert "PCD" in result.suggestedAcronyms
    assert "BHLS" in result.suggestedAcronyms
    assert "de" not in result.suggestedAcronyms
    assert "Troca" not in result.suggestedAcronyms


def _create_styled_template_xlsx() -> bytes:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Header"
    ws["A1"].font = Font(bold=True, size=14, color="FF0000", name="Arial")
    ws["A1"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    ws.merge_cells("A1:C1")
    ws.column_dimensions["A"].width = 15
    ws.row_dimensions[1].height = 30
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def test_preview_template():
    data = _create_template_xlsx()
    result = preview_template(data, "template.xlsx")
    assert len(result.sheets) >= 1
    assert any(c.value == "RDO Header" for c in result.sheets[0].cells)


def test_preview_template_styles():
    data = _create_styled_template_xlsx()
    result = preview_template(data, "template.xlsx")
    sheet = result.sheets[0]

    header_cell = next(c for c in sheet.cells if c.value == "Header")

    assert header_cell.style is not None
    assert header_cell.style.get("font-weight") == "bold"
    assert header_cell.style.get("font-size") == "14pt"
    assert header_cell.style.get("color") == "#FF0000"
    assert header_cell.style.get("font-family") == "Arial"
    assert header_cell.style.get("background-color") == "#FFFF00"
    assert header_cell.style.get("text-align") == "center"
    assert header_cell.style.get("vertical-align") == "center"
    assert "border" in header_cell.style.get("border-right", "")


def test_preview_template_merged_cells():
    data = _create_styled_template_xlsx()
    result = preview_template(data, "template.xlsx")
    sheet = result.sheets[0]

    assert len(sheet.merged) == 1
    m = sheet.merged[0]
    assert m["min_col"] == 1
    assert m["max_col"] == 3
    assert m["min_row"] == 1
    assert m["max_row"] == 1


def test_preview_template_dimensions():
    data = _create_styled_template_xlsx()
    result = preview_template(data, "template.xlsx")
    sheet = result.sheets[0]

    assert sheet.col_widths is not None
    assert sheet.col_widths.get("A") == 15
    assert sheet.row_heights is not None
    assert sheet.row_heights.get(1) == 30


def test_preview_template_auto_colors():
    from openpyxl.styles import PatternFill, Font

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Auto"
    ws["A1"].fill = PatternFill(fill_type=None)
    ws["A1"].font = Font(color=None)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    result = preview_template(buffer.read(), "auto.xlsx")
    cell = next(c for c in result.sheets[0].cells if c.value == "Auto")
    assert cell.style is not None
    assert "background-color" not in cell.style
    assert "color" not in cell.style
