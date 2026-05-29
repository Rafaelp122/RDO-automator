import base64
import io
from copy import deepcopy
from typing import BinaryIO

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage

from backend.logger import logger
from backend.schemas import CellData, ImageData, TemplatePreviewResponse, TemplateSheet
from backend.exceptions import InvalidFileExtension


class TemplateManager:
    def __init__(self, file: BinaryIO):
        self.file = file
        self.wb = None
        self.ws_template = None
        self.extracted_images = []

    def load(self):
        try:
            self.wb = load_workbook(self.file)
            self.ws_template = self.wb.active
            self.extracted_images = []
            if hasattr(self.ws_template, "_images"):
                for img in self.ws_template._images:
                    try:
                        img_bytes = img._data()
                        anchor = img.anchor
                        self.extracted_images.append({"bytes": img_bytes, "anchor": anchor})
                    except Exception as e:
                        logger.warning("Nao foi possivel ler uma imagem do template: %s", e)
                self.ws_template._images = []
        except Exception as e:
            logger.exception("Erro ao abrir template Excel")
            raise ValueError(f"O arquivo de template nao e um Excel (.xlsx) valido: {e}")

    def clone_worksheet(self, title: str):
        new_ws = self.wb.copy_worksheet(self.ws_template)
        new_ws.title = title
        for img_info in self.extracted_images:
            try:
                new_img = OpenpyxlImage(io.BytesIO(img_info["bytes"]))
                new_img.anchor = deepcopy(img_info["anchor"])
                new_ws.add_image(new_img)
            except Exception as e:
                logger.warning("Nao foi possivel injetar imagem na aba %s: %s", title, e)
        return new_ws

    def save_to_stream(self, stream: BinaryIO):
        if self.ws_template:
            self.wb.remove(self.ws_template)
        self.wb.save(stream)
        logger.info("Relatorio salvo no stream")


def preview_template(file_bytes: bytes, filename: str) -> TemplatePreviewResponse:
    if not filename.lower().endswith(".xlsx"):
        raise InvalidFileExtension("Template deve ser .xlsx")

    buffer = io.BytesIO(file_bytes)
    tm = TemplateManager(buffer)
    tm.load()

    all_sheets = []
    for ws in tm.wb.worksheets:
        cells = []
        for row in ws.iter_rows():
            for cell in row:
                font_info = None
                if cell.font:
                    font_info = {"bold": cell.font.bold, "size": cell.font.size}
                cells.append(CellData(
                    coord=cell.coordinate, row=cell.row, col=cell.column,
                    value=str(cell.value) if cell.value is not None else None,
                    font=font_info,
                ))

        images = []
        if hasattr(ws, "_images"):
            for img in ws._images:
                try:
                    img_bytes = img._data()
                    b64 = base64.b64encode(img_bytes).decode()
                    position = {}
                    if hasattr(img.anchor, "_from"):
                        position["col"] = img.anchor._from.col
                        position["row"] = img.anchor._from.row
                        position["colOff"] = img.anchor._from.colOff
                        position["rowOff"] = img.anchor._from.rowOff
                    images.append(ImageData(b64=f"data:image/png;base64,{b64}", position=position))
                except Exception as e:
                    logger.warning("Failed to extract image: %s", e)

        all_sheets.append(TemplateSheet(name=ws.title, cells=cells, images=images, merged=[]))

    logger.info("Preview template: %d sheets from %s", len(all_sheets), filename)
    return TemplatePreviewResponse(sheets=all_sheets)
