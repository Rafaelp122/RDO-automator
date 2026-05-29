import base64
import calendar
import io
import json
from copy import deepcopy
from datetime import datetime, timedelta
from typing import BinaryIO, Callable, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage

from logger import logger
from processor import TextProcessor
from schemas import (
    CellData, GenerateConfig, ImageData, SheetData,
    SourcePreviewResponse, TemplatePreviewResponse, TemplateSheet,
)
from exceptions import InvalidFileExtension, InvalidConfigError


# ---------------------------------------------------------------------------
# Extract
# ---------------------------------------------------------------------------

class ExcelLoader:
    def __init__(self, file: BinaryIO, header_row: int = 0):
        self.file = file
        self.header_row = header_row

    def load_all_sheets(self) -> dict[str, pd.DataFrame]:
        logger.info("Lendo dados do arquivo (header na linha %d)", self.header_row)
        abas = pd.read_excel(self.file, sheet_name=None, header=self.header_row)
        return self._normalize_dates(abas)

    def _normalize_dates(self, abas: dict[str, pd.DataFrame]) -> dict[str, pd.DataFrame]:
        for nome_aba, df in abas.items():
            col_data = self._find_date_column(df)
            if col_data:
                df[col_data] = pd.to_datetime(df[col_data], errors="coerce")
                df["_dia_aux"] = df[col_data].dt.day
            else:
                df["_dia_aux"] = None
        return abas

    @staticmethod
    def _find_date_column(df: pd.DataFrame) -> str:
        for col in df.columns:
            if "data" in str(col).lower():
                return col
        return ""


# ---------------------------------------------------------------------------
# Load (template)
# ---------------------------------------------------------------------------

class TemplateManager:
    def __init__(self, file: BinaryIO):
        self.file = file
        self.wb = None
        self.ws_template = None
        self.extracted_images = []

    def load(self):
        try:
            self.wb = load_workbook(self.file)
            self.ws_template = self.wb.active
            self.extracted_images = []
            if hasattr(self.ws_template, "_images"):
                for img in self.ws_template._images:
                    try:
                        img_bytes = img._data()
                        anchor = img.anchor
                        self.extracted_images.append({"bytes": img_bytes, "anchor": anchor})
                    except Exception as e:
                        logger.warning("Nao foi possivel ler uma imagem do template: %s", e)
                self.ws_template._images = []
        except Exception as e:
            logger.exception("Erro ao abrir template Excel")
            raise ValueError(f"O arquivo de template nao e um Excel (.xlsx) valido: {e}")

    def clone_worksheet(self, title: str):
        new_ws = self.wb.copy_worksheet(self.ws_template)
        new_ws.title = title
        for img_info in self.extracted_images:
            try:
                new_img = OpenpyxlImage(io.BytesIO(img_info["bytes"]))
                new_img.anchor = deepcopy(img_info["anchor"])
                new_ws.add_image(new_img)
            except Exception as e:
                logger.warning("Nao foi possivel injetar imagem na aba %s: %s", title, e)
        return new_ws

    def save_to_stream(self, stream: BinaryIO):
        if self.ws_template:
            self.wb.remove(self.ws_template)
        self.wb.save(stream)
        logger.info("Relatorio salvo no stream")


# ---------------------------------------------------------------------------
# Preview source
# ---------------------------------------------------------------------------

def preview_source(file_bytes: bytes, filename: str) -> SourcePreviewResponse:
    if not filename.lower().endswith((".xlsx", ".xls")):
        raise InvalidFileExtension("Arquivo de origem deve ser .xlsx ou .xls")

    buffer = io.BytesIO(file_bytes)
    loader = ExcelLoader(buffer)
    all_sheets = loader.load_all_sheets()

    sheets = []
    for name, df in all_sheets.items():
        cols = [str(c) for c in df.columns if c != "_dia_aux"]
        rows = df[cols].fillna("").head(20).values.tolist()
        string_rows = [[str(cell) for cell in row] for row in rows]
        sheets.append(SheetData(name=name, columns=cols, data=string_rows))

    logger.info("Preview source: %d sheets from %s", len(sheets), filename)
    return SourcePreviewResponse(sheets=sheets, filename=filename)


# ---------------------------------------------------------------------------
# Preview template
# ---------------------------------------------------------------------------

def preview_template(file_bytes: bytes, filename: str) -> TemplatePreviewResponse:
    if not filename.lower().endswith(".xlsx"):
        raise InvalidFileExtension("Template deve ser .xlsx")

    buffer = io.BytesIO(file_bytes)
    tm = TemplateManager(buffer)
    tm.load()

    all_sheets = []
    for ws in tm.wb.worksheets:
        cells = []
        for row in ws.iter_rows():
            for cell in row:
                font_info = None
                if cell.font:
                    font_info = {"bold": cell.font.bold, "size": cell.font.size}
                cells.append(CellData(
                    coord=cell.coordinate, row=cell.row, col=cell.column,
                    value=str(cell.value) if cell.value is not None else None,
                    font=font_info,
                ))

        images = []
        if hasattr(ws, "_images"):
            for img in ws._images:
                try:
                    img_bytes = img._data()
                    b64 = base64.b64encode(img_bytes).decode()
                    position = {}
                    if hasattr(img.anchor, "_from"):
                        position["col"] = img.anchor._from.col
                        position["row"] = img.anchor._from.row
                        position["colOff"] = img.anchor._from.colOff
                        position["rowOff"] = img.anchor._from.rowOff
                    images.append(ImageData(b64=f"data:image/png;base64,{b64}", position=position))
                except Exception as e:
                    logger.warning("Failed to extract image: %s", e)

        all_sheets.append(TemplateSheet(name=ws.title, cells=cells, images=images, merged=[]))

    logger.info("Preview template: %d sheets from %s", len(all_sheets), filename)
    return TemplatePreviewResponse(sheets=all_sheets)


# ---------------------------------------------------------------------------
# Generate report
# ---------------------------------------------------------------------------

def generate_report(
    source_bytes: bytes,
    template_bytes: bytes,
    source_filename: str,
    template_filename: str,
    config: dict | str,
) -> io.BytesIO:
    if not source_filename.lower().endswith((".xlsx", ".xls")):
        raise InvalidFileExtension("Arquivo de origem deve ser .xlsx ou .xls")
    if not template_filename.lower().endswith(".xlsx"):
        raise InvalidFileExtension("Template deve ser .xlsx")

    if isinstance(config, str):
        try:
            config_dict = json.loads(config)
        except json.JSONDecodeError as e:
            raise InvalidConfigError(f"JSON invalido: {e}")
    else:
        config_dict = config

    try:
        parsed = GenerateConfig(**config_dict)
    except Exception as e:
        raise InvalidConfigError(str(e))

    cfg = parsed.model_dump()
    contract = cfg["contract"]
    mappings = cfg["mappings"]
    list_separator = cfg.get("listSeparator", ", ")
    list_connector = cfg.get("listConnector", " e ")

    # --- inline ETL pipeline ---

    loader = ExcelLoader(io.BytesIO(source_bytes))
    abas_origem = loader.load_all_sheets()

    tm = TemplateManager(io.BytesIO(template_bytes))
    tm.load()

    ano = contract["ano"]
    mes = contract["mes"]
    _, ultimo_dia = calendar.monthrange(ano, mes)

    logger.info("Gerando relatorio: mes=%d/%d, %d mappings", mes, ano, len(mappings))

    for dia in range(1, ultimo_dia + 1):
        data_atual = datetime(ano, mes, dia)
        ws = tm.clone_worksheet(data_atual.strftime("%d-%m"))

        for mapping in mappings:
            celula = mapping.get("templateCell", "")
            if not celula:
                continue

            format_template = mapping.get("formatTemplate", "")
            source_columns = mapping.get("sourceColumns", [])

            combined_values = {}
            for sheet_name, df in abas_origem.items():
                if "_dia_aux" not in df.columns:
                    continue
                filtro = df[df["_dia_aux"] == dia]
                if filtro.empty:
                    continue
                for col in source_columns:
                    if col in filtro.columns:
                        vals = filtro[col].dropna().unique().tolist()
                        vals = [v for v in vals if str(v).strip() and str(v).lower() != "nan"]
                        if vals:
                            if col not in combined_values:
                                combined_values[col] = []
                            combined_values[col].extend(vals)

            if not combined_values:
                ws[celula] = None
                continue

            for col in combined_values:
                combined_values[col] = list(dict.fromkeys(combined_values[col]))

            resumo = TextProcessor.formatar_resumo(combined_values, format_template, list_separator, list_connector)
            ws[celula] = resumo if resumo else None

    output = io.BytesIO()
    tm.save_to_stream(output)
    output.seek(0)
    logger.info("Relatorio gerado com sucesso")
    return output
