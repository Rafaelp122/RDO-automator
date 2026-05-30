import io
import json

import pandas as pd
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from src.main import app

client = TestClient(app)


def test_full_flow():
    source_buffer = io.BytesIO()
    with pd.ExcelWriter(source_buffer, engine="openpyxl") as writer:
        df1 = pd.DataFrame(
            {
                "Data": ["2026-01-05", "2026-01-10", "2026-01-15", "2026-01-20"],
                "Serviço": ["Pintura", "Concretagem", "Escavacao", "Armação"],
                "Bairro": ["Centro", "Vista Alegre", "Centro", "Leste"],
            }
        )
        df1.to_excel(writer, sheet_name="Relatorio_Geral", index=False)
    source_buffer.seek(0)

    template_buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "RDO - RELATORIO DIARIO DE OBRA"
    ws["B3"] = ""
    ws["B5"] = ""
    wb.save(template_buffer)
    template_buffer.seek(0)

    config = {
        "contract": {"start_date": "2026-01-01", "prazo_dias": 180, "mes": 1, "ano": 2026},
        "mappings": [
            {
                "formatTemplate": "Servicos Realizados: {Serviço}. No Bairro: {Bairro}",
                "templateCell": "B3",
                "sourceColumns": ["Serviço", "Bairro"],
            }
        ],
        "listSeparator": ", ",
        "listConnector": " e ",
    }

    response = client.post(
        "/api/generate",
        files={
            "source": (
                "medicao.xlsx",
                source_buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            "template": (
                "template.xlsx",
                template_buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
        data={"config": json.dumps(config)},
    )

    assert response.status_code == 200
    assert "spreadsheetml" in response.headers["content-type"]

    output = io.BytesIO(response.content)
    wb_out = load_workbook(output)
    assert "05-01" in wb_out.sheetnames
    ws_out = wb_out["05-01"]
    cell_val = ws_out["B3"].value
    assert cell_val is not None
    assert "Servicos Realizados:" in str(cell_val)
    assert "Pintura" in str(cell_val)
    assert "Centro" in str(cell_val)
