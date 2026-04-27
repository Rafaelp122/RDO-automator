import pytest
import pandas as pd
from openpyxl import Workbook
from pathlib import Path
from unittest.mock import patch
from src.app.core.config_models import ReportConfig
from src.app.core import constants

@pytest.fixture(autouse=True)
def mock_constants(tmp_path, monkeypatch):
    """Sobrescreve os diretórios globais para apontar para a pasta temporária do pytest."""
    test_root = tmp_path / "app_root"
    test_root.mkdir()
    
    log_dir = test_root / "logs"
    input_dir = test_root / "data" / "input"
    output_dir = test_root / "data" / "output"
    
    for d in [log_dir, input_dir, output_dir]:
        d.mkdir(parents=True, exist_ok=True)
    
    # Monkeypatch para garantir que qualquer módulo que já importou as constantes veja as novas
    monkeypatch.setattr(constants, "LOG_DIR", log_dir)
    monkeypatch.setattr(constants, "INPUT_DIR", input_dir)
    monkeypatch.setattr(constants, "OUTPUT_DIR", output_dir)
    monkeypatch.setattr(constants, "APP_LOG_PATH", log_dir / "app.log")
    
    # Também patchear diretamente onde são usados para garantir
    monkeypatch.setattr("src.app.core.report_service.OUTPUT_DIR", output_dir)

    yield

@pytest.fixture
def temp_excel_dir(tmp_path):
    """Fixture para criar diretórios de entrada e saída temporários."""
    input_dir = tmp_path / "input_manual"
    output_dir = tmp_path / "output_manual"
    input_dir.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)
    return input_dir, output_dir

@pytest.fixture
def sample_data_excel(temp_excel_dir):
    """Cria uma planilha de dados de origem válida para testes."""
    input_dir, _ = temp_excel_dir
    path = input_dir / "origem.xlsx"
    
    df1 = pd.DataFrame({
        'Data': [pd.Timestamp('2026-04-01'), pd.Timestamp('2026-04-02')],
        'Servico': ['Escavação', 'Aterro'],
        'Local': ['Rua A', 'Rua B']
    })
    
    df2 = pd.DataFrame({
        'Data': [pd.Timestamp('2026-04-01')],
        'Equipamento': ['Retroescavadeira'],
        'Horas': [8]
    })
    
    with pd.ExcelWriter(path) as writer:
        df1.to_excel(writer, sheet_name='Obras', index=False)
        df2.to_excel(writer, sheet_name='Equipamentos', index=False)
    
    return path

@pytest.fixture
def sample_template_excel(temp_excel_dir):
    """Cria um arquivo de template .xlsx válido."""
    input_dir, _ = temp_excel_dir
    path = input_dir / "template.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    ws["A1"] = "DATA ATUAL:"
    ws["B10"] = "RESUMO SERVIÇOS:"
    wb.save(path)
    return path

@pytest.fixture
def valid_config(sample_data_excel, sample_template_excel):
    """Retorna um objeto ReportConfig populado com caminhos válidos."""
    config = ReportConfig()
    config.projeto.ano = 2026
    config.projeto.mes = 4
    config.arquivos.dados_origem = str(sample_data_excel)
    config.arquivos.user_template = str(sample_template_excel)
    config.arquivos.template_ativo = str(sample_template_excel)
    config.posicoes.celula_data_atual = "B1"
    config.extracao.colunas = ["Data", "Servico", "Local"]
    config.extracao.formato_final = "{Servico} em {Local}"
    config.mapeamento = {"Obras": "B10"}
    return config
