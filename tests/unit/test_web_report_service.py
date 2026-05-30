import io
import pandas as pd
from openpyxl import Workbook, load_workbook
from src.excel import ReportGenerator


def _create_source_xlsx() -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df = pd.DataFrame({
            "Data": ["2026-01-02"],
            "Atividade": ["Concretagem"],
        })
        df.to_excel(writer, sheet_name="Obra", index=False)
    buffer.seek(0)
    return buffer.read()


def _create_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "RDO"
    ws["B3"] = ""
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def test_generate_report():
    source = _create_source_xlsx()
    template = _create_template_xlsx()

    config = {
        "contract": {"start_date": "2026-01-01", "prazo_dias": 30, "mes": 1, "ano": 2026},
        "mappings": [{"formatTemplate": "{Atividade}", "templateCell": "B3", "sourceColumns": ["Atividade"]}],
        "listSeparator": ", ",
        "listConnector": " e ",
    }

    output = ReportGenerator(
        source, template, "medicao.xlsx", "template.xlsx", config
    ).generate()
    output.seek(0)

    wb = load_workbook(output)
    sheets = wb.sheetnames
    assert len(sheets) >= 30
    assert "05-01" in sheets
