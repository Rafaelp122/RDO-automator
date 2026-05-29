import io
import pandas as pd
from openpyxl import Workbook
from backend.excel import preview_source, preview_template


def _create_source_xlsx() -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        pd.DataFrame({"Data": ["2026-01-01"], "Atividade": ["Escavacao"]}).to_excel(writer, sheet_name="Obra", index=False)
    buffer.seek(0)
    return buffer.read()


def _create_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "RDO Header"
    ws["B3"] = "Servicos:"
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def test_preview_source():
    data = _create_source_xlsx()
    result = preview_source(data, "medicao.xlsx")
    assert len(result.sheets) == 1
    assert result.sheets[0].name == "Obra"
    assert "Atividade" in result.sheets[0].columns


def test_preview_template():
    data = _create_template_xlsx()
    result = preview_template(data, "template.xlsx")
    assert len(result.sheets) >= 1
    assert any(c.value == "RDO Header" for c in result.sheets[0].cells)
