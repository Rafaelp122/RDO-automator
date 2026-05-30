import io

from openpyxl import Workbook

from src.excel import TemplateManager


def _create_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "RDO - RELATORIO DIARIO DE OBRA"
    ws["B3"] = "Servicos:"
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def test_load_from_bytesio():
    data = _create_template_xlsx()
    buffer = io.BytesIO(data)
    tm = TemplateManager(buffer)
    assert tm.wb is not None
    assert tm.ws_template is not None
    assert tm.ws_template["A1"].value == "RDO - RELATORIO DIARIO DE OBRA"


def test_clone_worksheet():
    data = _create_template_xlsx()
    buffer = io.BytesIO(data)
    tm = TemplateManager(buffer)
    ws = tm.clone_worksheet("01-01")
    assert ws.title == "01-01"
    assert ws["A1"].value == "RDO - RELATORIO DIARIO DE OBRA"
    assert len(tm.wb.sheetnames) == 2


def test_save_to_bytesio():
    data = _create_template_xlsx()
    buffer = io.BytesIO(data)
    tm = TemplateManager(buffer)
    tm.clone_worksheet("01-01")
    output = io.BytesIO()
    tm.save_to_stream(output)
    output.seek(0)
    assert output.read(4) == b"PK\x03\x04"
